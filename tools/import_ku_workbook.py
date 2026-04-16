#!/usr/bin/env python3
"""
Regenerate KU student week workshop data from Excel.

Source: backend/database/source/ku-workshops-schedule.xlsx
Sheets used:
  - «مواعيد الورش» — 100 workshop cells (5 tracks × 4 slots × 5 days)
  - «ارقام المحاضرين» — facilitator phones

Outputs:
  - backend/database/data/ku_student_week_workshops_schedule_full.json
  - frontend/src/app/core/data/ku_student_week_workshops_schedule_full.json (identical)
  - backend/database/data/ku_student_week_workshops.php
  - backend/database/data/ku_student_week_facilitators.json

Run from repo root:
  tools/.venv/bin/python tools/import_ku_workbook.py

Requires openpyxl (tools/.venv).
"""
from __future__ import annotations

import difflib
import json
import re
import subprocess
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "backend/database/source/ku-workshops-schedule.xlsx"
LEGACY_PHP = ROOT / "backend/database/data/ku_student_week_workshops.php"
OUT_SCHEDULE_BE = ROOT / "backend/database/data/ku_student_week_workshops_schedule_full.json"
OUT_SCHEDULE_FE = ROOT / "frontend/src/app/core/data/ku_student_week_workshops_schedule_full.json"
OUT_PHP = ROOT / "backend/database/data/ku_student_week_workshops.php"
OUT_FAC = ROOT / "backend/database/data/ku_student_week_facilitators.json"

IMAGES = [
    "https://images.unsplash.com/photo-1524178232363-1fb2b075b655?auto=format&fit=crop&w=800&q=80",
    "https://images.unsplash.com/photo-1517245386807-bb43f82c33c4?auto=format&fit=crop&w=800&q=80",
    "https://images.unsplash.com/photo-1522071820081-009f0129c71c?auto=format&fit=crop&w=800&q=80",
]

# Normalize presenter strings between schedule sheet and legacy seed / phone sheet.
PRESENTER_ALIASES = {
    "علي عادل": "علي الأنصاري",
    "أحمد سمير": "أحمد سمير",
    "سليمان المراغى": "سليمان المراغي",
    "د.بسام الجزاف": "د. بسام الجزاف",
    "د.بسام الجزاف ": "د. بسام الجزاف",
    "الحكم خالد الشمري": "خالد الشمري",
    "م. زينب الغضبان": "زينب الغضبان",
    "منيرة النخلان": "منيرة النخيلان",
    "إبراهيم السماعيل": "المحامي إبراهيم السماعيل",
    "د. بسام الجزاف": "د. بسام الجزاف",
}

# When Arabic title + presenter (after aliases) still diverge from legacy, use these English titles.
TITLE_EN_OVERRIDES: dict[str, str] = {
    "ms-w-002": "Will AI take your job—or help you land one?",
    "ms-w-018": "The first step",
    "ms-w-019": "Photography fundamentals",
    "ms-w-020": "Basketball referees",
    "ms-w-023": "Software used in graphic design",
    "ms-w-025": "Success mindset",
    "ms-w-030": "Workplace conflict resolution",
    "ms-w-040": "Refereeing fundamentals",
    "ms-w-049": "Planning science and tools for work quality",
    "ms-w-055": "Welcome income",
    "ms-w-059": "Visual analysis and extracting ideas",
    "ms-w-060": "Video tech and the IRS referee system",
    "ms-w-062": "Releasing negative emotions",
    "ms-w-064": "Local and international refereeing",
    "ms-w-070": "From colleague to partner",
    "ms-w-072": "Business model canvas workshop",
    "ms-w-073": "Managing emotions at work",
    "ms-w-077": "SMART goal setting",
    "ms-w-079": "Smartphone video basics (while filming)",
    "ms-w-080": "Growing profit through social media programs",
    "ms-w-084": "Using AI in digital art",
    "ms-w-085": "From graduate to employee",
    "ms-w-090": "Why become a referee",
    "ms-w-100": "From a shot to a professional story",
}


def norm_ws(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def canon_presenter(p: str) -> str:
    p = norm_ws(p)
    return PRESENTER_ALIASES.get(p, p)


def looks_like_presenter_line(s: str) -> bool:
    """Heuristic: last line(s) of a cell are the facilitator name, not a title continuation."""
    s = norm_ws(s)
    if not s or len(s) > 90:
        return False
    low = s.lower()
    if low.startswith("http"):
        return False
    if s.startswith(("برامج", "لضمان", "(قبل", "(أثناء", "(بعد")):
        return False
    if any(
        s.startswith(p)
        for p in ("د.", "د .", "المحامي", "أ.", "أ .", "م.", "م .", "الحكم", "شهد ", "هيا ")
    ):
        return True
    if " و " in s and len(s.split()) <= 10:
        return True
    toks = s.split()
    if 1 <= len(toks) <= 5 and "برنامج" not in s and "التصوير" not in s:
        return True
    if "شهد العطار" in s or ("حمزه" in s and "شهد" in s):
        return True
    return False


def split_title_presenter(cell) -> tuple[str, str]:
    if cell is None:
        return "", ""
    s = str(cell).replace("\r\n", "\n").replace("\r", "\n").strip()
    parts = [norm_ws(p) for p in s.split("\n") if norm_ws(p)]
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    for i in range(len(parts) - 1, -1, -1):
        if looks_like_presenter_line(parts[i]):
            title = norm_ws(" ".join(parts[:i]))
            return title, parts[i]
    return norm_ws(" ".join(parts[:-1])), parts[-1]


def parse_time_cell(a_str: str) -> str | None:
    if not a_str:
        return None
    s = str(a_str).strip()
    m = re.match(r"(\d{1,2}):(\d{2})", s)
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    if "م" in s and "ص" not in s and h != 12:
        h += 12
    return f"{h:02d}:{mi:02d}"


def parse_date_from_day(cell) -> str | None:
    s = str(cell)
    m = re.search(r"(\d{2})-(\d{2})", s)
    if not m:
        return None
    return f"2026-{m.group(2)}-{m.group(1)}"


def load_legacy_pack() -> dict:
    cmd = [
        "php",
        "-r",
        f'echo json_encode(require {json.dumps(str(LEGACY_PHP))}, JSON_UNESCAPED_UNICODE);',
    ]
    raw = subprocess.check_output(cmd, cwd=str(ROOT / "backend"))
    return json.loads(raw.decode("utf-8"))


def parse_schedule(ws) -> list[dict]:
    current_date: str | None = None
    grid: list[dict] = []
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        a_str = str(a).strip() if a is not None else ""
        if "اليوم" in a_str and re.search(r"\d{2}-\d{2}", a_str):
            current_date = parse_date_from_day(a_str)
            continue
        t = parse_time_cell(a_str)
        if not t or not current_date:
            continue
        has_cells = any(ws.cell(r, c).value for c in range(2, 7))
        if not has_cells:
            continue
        for track in range(5):
            cell = ws.cell(r, track + 2).value
            title_ar, presenter_ar = split_title_presenter(cell)
            grid.append(
                {
                    "date": current_date,
                    "time": t,
                    "track": track + 1,
                    "title_ar": title_ar,
                    "presenter_ar": presenter_ar,
                }
            )
    return grid


def parse_facilitators(ws) -> list[dict]:
    out: list[dict] = []
    for r in range(2, ws.max_row + 1):
        phone_raw = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        joined = ws.cell(r, 3).value
        if not name or not str(name).strip():
            continue
        name_s = norm_ws(str(name))
        if "اعتذرت" in name_s:
            continue
        phone = None
        if phone_raw is not None and str(phone_raw).strip() != "":
            try:
                n = int(round(float(phone_raw)))
                s = str(n)
                if len(s) == 8:
                    phone = "+965" + s
                elif len(s) == 11 and s.startswith("965"):
                    phone = "+" + s
                else:
                    phone = "+" + s
            except (TypeError, ValueError):
                phone = None
        on_platform = None
        if joined is not None:
            js = str(joined).strip()
            if js in ("نعم", "yes", "Yes", "TRUE", "true", "1"):
                on_platform = True
            elif js in ("لا", "no", "No", "FALSE", "false", "0"):
                on_platform = False
        out.append(
            {
                "name_ar": canon_presenter(name_s),
                "phone": phone,
                "on_platform": on_platform,
            }
        )
    return out


def pick_title_en(
    slug: str,
    new_title: str,
    new_pres: str,
    legacy_en: str,
    old_title: str,
    old_pres: str,
) -> str:
    if slug in TITLE_EN_OVERRIDES:
        return TITLE_EN_OVERRIDES[slug]
    r1 = difflib.SequenceMatcher(None, norm_ws(new_title), norm_ws(old_title)).ratio()
    r2 = difflib.SequenceMatcher(
        None, canon_presenter(new_pres), canon_presenter(old_pres)
    ).ratio()
    if r1 >= 0.82 and r2 >= 0.82:
        return legacy_en
    return TITLE_EN_OVERRIDES.get(slug, legacy_en)


def php_quote(s: str) -> str:
    return "'" + s.replace("\\", "\\\\").replace("'", "\\'") + "'"


def emit_php(rows: list[tuple], title_en: dict[str, str]) -> str:
    lines = [
        "<?php",
        "",
        "/**",
        " * April 2026 KU workshop week — generated by tools/import_ku_workbook.py from",
        " * database/source/ku-workshops-schedule.xlsx (tab «مواعيد الورش»).",
        " *",
        " * @see database/data/ku_student_week_workshops_schedule_full.json (keep in sync)",
        " */",
        "return [",
        "    'images' => [",
    ]
    for img in IMAGES:
        lines.append(f"        {php_quote(img)},")
    lines.append("    ],")
    lines.append("    /** [slug, title_ar, presenter, 'Y-m-d', 'H:i', featured] */")
    lines.append("    'rows' => [")
    for slug, title, pres, d, tim, feat in rows:
        fe = "true" if feat else "false"
        lines.append(
            f"        [{php_quote(slug)}, {php_quote(title)}, {php_quote(pres)}, {php_quote(d)}, {php_quote(tim)}, {fe}],"
        )
    lines.append("    ],")
    lines.append("    'title_en_by_slug' => [")
    for slug in sorted(title_en.keys(), key=lambda x: int(x.split("-")[-1])):
        lines.append(f"        {php_quote(slug)} => {php_quote(title_en[slug])},")
    lines.append("    ],")
    lines.append("];")
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    if not XLSX.exists():
        print("Missing", XLSX, file=sys.stderr)
        return 1

    legacy = load_legacy_pack()
    old_rows = legacy["rows"]
    old_title_en: dict[str, str] = legacy["title_en_by_slug"]

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    grid = parse_schedule(wb["مواعيد الورش"])
    if len(grid) != 100:
        print(f"Expected 100 workshops, got {len(grid)}", file=sys.stderr)
        return 1

    workshops_json = []
    php_rows: list[tuple] = []
    title_en_out: dict[str, str] = {}

    for i, g in enumerate(grid):
        n = i + 1
        slug = f"ms-w-{n:03d}"
        _, _, _, _, _, feat = old_rows[i]
        title_ar = g["title_ar"]
        presenter_ar = canon_presenter(g["presenter_ar"])
        php_rows.append(
            (slug, title_ar, presenter_ar, g["date"], g["time"], bool(feat))
        )
        workshops_json.append(
            {
                "n": n,
                "date": g["date"],
                "time": g["time"],
                "track": g["track"],
                "title_ar": title_ar,
                "presenter_ar": presenter_ar,
            }
        )
        oslug, ot, op, _, _, _ = old_rows[i]
        assert oslug == slug
        title_en_out[slug] = pick_title_en(
            slug, title_ar, presenter_ar, old_title_en[slug], ot, op
        )

    schedule_obj = {
        "meta": {
            "locale": "ar",
            "date_range": "2026-04-26 — 2026-04-30",
            "slots_per_day": ["09:30", "10:30", "11:30", "12:30"],
            "parallel_tracks": 5,
            "total_workshops": 100,
        },
        "workshops": workshops_json,
    }

    fac = parse_facilitators(wb["ارقام المحاضرين"])

    json_text = json.dumps(schedule_obj, ensure_ascii=False, indent=2) + "\n"
    OUT_SCHEDULE_BE.write_text(json_text, encoding="utf-8")
    OUT_SCHEDULE_FE.write_text(json_text, encoding="utf-8")

    OUT_PHP.write_text(emit_php(php_rows, title_en_out), encoding="utf-8")
    OUT_FAC.write_text(
        json.dumps(fac, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    print("Wrote:", OUT_SCHEDULE_BE)
    print("Wrote:", OUT_SCHEDULE_FE)
    print("Wrote:", OUT_PHP)
    print("Wrote:", OUT_FAC)
    print("Facilitators:", len(fac))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
